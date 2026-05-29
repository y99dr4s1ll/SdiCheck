"""
Controllo mensile fatture - COESA
Interfaccia grafica con tkinter (incluso in Python, nessuna installazione aggiuntiva)
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import subprocess, sys, os


# ─── LOGICA ───────────────────────────────────────────────────────────────────

def carica_fatture(path):
    return pd.read_excel(path, dtype={
        "Cod. documento": str, "Sz.": int, "Nr.doc.": int,
        "Tipo doc.": str, "Ragione sociale anagrafica": str,
        "IVA": str, "Alias": str, "Cod. articolo": str,
        "Descrizione articolo": str, "Um 1": str,
        "Codice CIG": str, "Codice CUP": str, "Macrocategoria": str,
    }, parse_dates=["Data doc."])


def carica_csv(path):
    for sep in [";", ","]:
        try:
            df = pd.read_csv(path, sep=sep, dtype=str, encoding="utf-8-sig")
            if df.shape[1] > 1:
                return df
        except Exception:
            continue
    raise ValueError(f"Impossibile leggere il CSV: {path}")


def elabora(df_fat, df_csv, codice_sz, data_filtro):
    def to_num(series):
        return pd.to_numeric(series.astype(str).str.replace(",", ".", regex=False), errors="coerce").fillna(0)

    # Escludi C-ORDINE-PA
    df_fat = df_fat[df_fat["Cod. documento"] != "C-ORDINE-PA"].copy()

    nr = df_fat["Nr.doc."].astype(str).str.zfill(3)
    sz = df_fat["Sz."].astype(str).str.zfill(2)
    df_fat["Numero Fattura"] = "'" + nr + "/" + sz + "'"

    # Supporto codici Sz. multipli separati da virgola (es. "0,1" o "2")
    codici = [int(c.strip()) for c in str(codice_sz).split(",") if c.strip() != ""]
    df_fat = df_fat[df_fat["Sz."].isin(codici)].copy()
    df_fat = df_fat[df_fat["Data doc."] > pd.Timestamp(data_filtro)].copy()

    # C-NOTACR-PA → Tot. imponibile e Tot. Iva negativi
    nc_mask = df_fat["Cod. documento"] == "C-NOTACR-PA"
    for col in ["Tot. imponibile", "Tot. Iva"]:
        if col in df_fat.columns:
            df_fat[col] = to_num(df_fat[col])
            df_fat.loc[nc_mask, col] = -df_fat.loc[nc_mask, col].abs()

    agg_kwargs = {"Imp. FattureEmesse": ("Tot. imponibile", "first"),
                  "IVA FattureEmesse":  ("Tot. Iva", "first")}
    grouped = (
        df_fat
        .groupby(["Numero Fattura", "Data doc.", "Cli./For.", "Ragione sociale anagrafica"], as_index=False)
        .agg(**agg_kwargs)
    )

    df_csv = df_csv.rename(columns={"Numero fattura / Documento": "Numero Fattura CSV"})

    # Nota di credito → importi negativi
    imp_col = "Imponibile/Importo (totale in euro)"
    iva_col = "Imposta (totale in euro)"
    if imp_col in df_csv.columns:
        df_csv[imp_col] = to_num(df_csv[imp_col])
    if iva_col in df_csv.columns:
        df_csv[iva_col] = to_num(df_csv[iva_col])

    tipo_col = "Tipo documento" if "Tipo documento" in df_csv.columns else None
    if tipo_col:
        is_nc = df_csv[tipo_col].astype(str).str.strip().str.lower().str.contains("nota di credito")
        if imp_col in df_csv.columns:
            df_csv.loc[is_nc, imp_col] = -df_csv.loc[is_nc, imp_col].abs()
        if iva_col in df_csv.columns:
            df_csv.loc[is_nc, iva_col] = -df_csv.loc[is_nc, iva_col].abs()

    # Stato consegna SDI (per gestire riemissioni con stesso numero)
    stato_sdi_col = "Fatture consegnate" if "Fatture consegnate" in df_csv.columns else None

    base_cols = ["Numero Fattura CSV", "Codice fiscale cliente", imp_col, iva_col, "Sdi/file"]
    rename_map = {imp_col: "Imp. CSV", iva_col: "IVA CSV"}
    if stato_sdi_col:
        base_cols.append(stato_sdi_col)
        rename_map[stato_sdi_col] = "Stato SDI"
    df_csv_sel = df_csv[base_cols].rename(columns=rename_map)

    # Riemissioni con stesso numero (es. fattura rifiutata e poi riemessa/accettata):
    # NON si sommano — si tiene UNA sola riga per numero fattura, preferendo quella accettata
    # e scartando quella rifiutata.
    n_dup = int(df_csv_sel.duplicated(subset="Numero Fattura CSV").sum())
    if "Stato SDI" in df_csv_sel.columns:
        # rank 0 = non rifiutata (preferita), 1 = rifiutata (in fondo)
        df_csv_sel["__rank"] = (
            df_csv_sel["Stato SDI"].astype(str).str.lower().str.contains("rifiutat").astype(int)
        )
        df_csv_sel = df_csv_sel.sort_values("__rank").drop(columns="__rank")
    df_csv_sel = df_csv_sel.drop_duplicates(subset="Numero Fattura CSV", keep="first")

    merged = pd.merge(
        grouped,
        df_csv_sel,
        left_on="Numero Fattura", right_on="Numero Fattura CSV", how="outer"
    )

    def stato(row):
        in_fat = pd.notna(row.get("Numero Fattura")) and row.get("Numero Fattura", "") != ""
        in_csv = pd.notna(row.get("Numero Fattura CSV")) and row.get("Numero Fattura CSV", "") != ""
        if in_fat and in_csv:   return "✔ OK"
        elif in_fat:            return "⚠ Solo FattureEmesse"
        else:                   return "⚠ Solo SDI"

    merged["Stato"] = merged.apply(stato, axis=1)
    # Colonna delta: Imp. FattureEmesse - Imp. CSV (arrotondato a 0,001)
    merged["Delta"] = (
        pd.to_numeric(merged["Imp. FattureEmesse"], errors="coerce").fillna(0) -
        pd.to_numeric(merged["Imp. CSV"], errors="coerce").fillna(0)
    ).round(3)

    cols = ["Stato", "Numero Fattura", "Data doc.", "Cli./For.", "Ragione sociale anagrafica",
            "Imp. FattureEmesse", "IVA FattureEmesse", "Imp. CSV", "IVA CSV", "Delta",
            "Numero Fattura CSV", "Codice fiscale cliente", "Sdi/file", "Stato SDI"]
    result = merged[[c for c in cols if c in merged.columns]]

    # ── Totali per il foglio Riepilogo (derivati dagli stessi valori del dettaglio) ──
    imp_csv = to_num(merged.get("Imp. CSV", pd.Series(dtype=float))).sum()
    iva_csv = to_num(merged.get("IVA CSV", pd.Series(dtype=float))).sum()
    imp_fat = pd.to_numeric(grouped["Imp. FattureEmesse"], errors="coerce").fillna(0).sum()
    iva_fat = pd.to_numeric(grouped["IVA FattureEmesse"], errors="coerce").fillna(0).sum()

    totali = {
        "Imponibile CSV":         imp_csv,
        "IVA CSV":                iva_csv,
        "Imponibile FattureEmesse": imp_fat,
        "IVA FattureEmesse":      iva_fat,
        "Delta Imponibile":       imp_fat - imp_csv,
        "Delta IVA":              iva_fat - iva_csv,
        "Duplicati CSV":          int(n_dup),
    }
    return result, totali


def formatta_excel(path, df, totali):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Controllo")

    wb = load_workbook(path)
    ws = wb["Controllo"]
    hf  = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    hfont = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    ok_f  = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
    warn_f= PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6")
    bfont = Font(name="Calibri", size=10)
    thin  = Side(style="thin", color="BFBFBF")
    brd   = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = hf; cell.font = hfont
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = brd

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        fill = ok_f if "OK" in str(row[0].value or "") else warn_f
        for cell in row:
            cell.font = bfont; cell.border = brd
            cell.alignment = Alignment(vertical="center"); cell.fill = fill

    for col_cells in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 40)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # ── Foglio Riepilogo ──
    ws2 = wb.create_sheet("Riepilogo")
    hf2   = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
    hfont2= Font(bold=True, color="FFFFFF", name="Calibri", size=10)
    ok_f2 = PatternFill("solid", start_color="E2EFDA", end_color="E2EFDA")
    warn_f2=PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6")
    bfont2= Font(name="Calibri", size=10)
    num_fmt = '#,##0.00'

    headers = ["Voce", "CSV / SDI", "FattureEmesse", "Delta"]
    for c, h in enumerate(headers, 1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.fill = hf2; cell.font = hfont2
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = brd

    rows_data = [
        ("Imponibile", totali["Imponibile CSV"], totali["Imponibile FattureEmesse"], totali["Delta Imponibile"]),
        ("IVA",        totali["IVA CSV"],        totali["IVA FattureEmesse"],        totali["Delta IVA"]),
        ("Totale",
         totali["Imponibile CSV"] + totali["IVA CSV"],
         totali["Imponibile FattureEmesse"] + totali["IVA FattureEmesse"],
         totali["Delta Imponibile"] + totali["Delta IVA"]),
    ]

    for r, (voce, csv_val, fat_val, delta) in enumerate(rows_data, 2):
        is_warn = abs(delta) > 0.01
        row_fill = warn_f2 if is_warn else ok_f2
        for c, val in enumerate([voce, csv_val, fat_val, delta], 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.font = Font(name="Calibri", size=10, bold=(voce == "Totale"))
            cell.border = brd
            cell.alignment = Alignment(horizontal="right" if c > 1 else "left", vertical="center")
            cell.fill = row_fill
            if c > 1:
                cell.number_format = num_fmt
            # Delta in rosso se diverso da zero
            if c == 4 and is_warn:
                cell.font = Font(name="Calibri", size=10, bold=(voce == "Totale"), color="C00000")

    for col_cells in ws2.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col_cells)
        ws2.column_dimensions[get_column_letter(col_cells[0].column)].width = max_len + 6

    ws2.sheet_view.showGridLines = False

    wb.save(path)


# ─── GUI ──────────────────────────────────────────────────────────────────────

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Controllo Fatture COESA")
        self.resizable(False, False)
        self.configure(bg="#F0F4F8")
        self._build_ui()
        self._center()

    def _center(self):
        self.update_idletasks()
        w, h = self.winfo_width(), self.winfo_height()
        sw, sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"+{(sw-w)//2}+{(sh-h)//2}")

    def _build_ui(self):
        BLUE   = "#1F4E79"
        LBLUE  = "#2E75B6"
        BG     = "#F0F4F8"
        WHITE  = "#FFFFFF"
        GRAY   = "#E2E8F0"

        # ── Titolo ──
        hdr = tk.Frame(self, bg=BLUE, pady=14)
        hdr.pack(fill="x")
        tk.Label(hdr, text="Controllo Mensile Fatture", font=("Calibri", 16, "bold"),
                 bg=BLUE, fg=WHITE).pack()
        tk.Label(hdr, text="COESA — incrocio FattureEmesse ↔ SDI", font=("Calibri", 9),
                 bg=BLUE, fg="#A8C4E0").pack()

        body = tk.Frame(self, bg=BG, padx=28, pady=20)
        body.pack(fill="both")

        def section(parent, title):
            f = tk.LabelFrame(parent, text=f"  {title}  ", font=("Calibri", 9, "bold"),
                              bg=BG, fg=LBLUE, bd=1, relief="groove", padx=12, pady=10)
            f.pack(fill="x", pady=(0, 14))
            return f

        def file_row(parent, label, var, filetypes, row):
            tk.Label(parent, text=label, font=("Calibri", 9), bg=BG, width=18, anchor="w"
                     ).grid(row=row, column=0, sticky="w", pady=4)
            e = tk.Entry(parent, textvariable=var, font=("Calibri", 9), width=46,
                         bg=WHITE, relief="solid", bd=1)
            e.grid(row=row, column=1, padx=(6, 6))
            tk.Button(parent, text="Sfoglia…", font=("Calibri", 8), bg=GRAY, relief="flat",
                      cursor="hand2",
                      command=lambda v=var, ft=filetypes: self._browse(v, ft)
                      ).grid(row=row, column=2)

        # ── File ──
        self.v_fatture = tk.StringVar()
        self.v_csv     = tk.StringVar()
        self.v_output  = tk.StringVar()

        sec1 = section(body, "File di input / output")
        file_row(sec1, "FattureEmesse (.xlsx)", self.v_fatture,
                 [("Excel", "*.xlsx")], 0)
        file_row(sec1, "Clienti SDI (.csv)",   self.v_csv,
                 [("CSV", "*.csv")],    1)
        file_row(sec1, "Output (.xlsx)",        self.v_output,
                 [("Excel", "*.xlsx")], 2)

        # ── Parametri ──
        sec2 = section(body, "Parametri elaborazione")
        sec2.columnconfigure(1, weight=1)

        def lbl(text, row):
            tk.Label(sec2, text=text, font=("Calibri", 9), bg=BG, anchor="w"
                     ).grid(row=row, column=0, sticky="w", pady=4)

        self.v_data   = tk.StringVar(value="2026-03-31")
        self.v_sz     = tk.StringVar(value="2")

        lbl("Data filtro (esclusa):", 0)
        tk.Entry(sec2, textvariable=self.v_data, font=("Calibri", 9), width=14,
                 bg=WHITE, relief="solid", bd=1).grid(row=0, column=1, sticky="w", padx=6)
        tk.Label(sec2, text="es. 2026-03-31  →  prende date > fine mese precedente",
                 font=("Calibri", 8), bg=BG, fg="#718096").grid(row=0, column=2, sticky="w")

        lbl("Codice Sz.:", 1)
        tk.Entry(sec2, textvariable=self.v_sz, font=("Calibri", 9), width=6,
                 bg=WHITE, relief="solid", bd=1).grid(row=1, column=1, sticky="w", padx=6)

        # ── Log ──
        sec3 = section(body, "Log elaborazione")
        self.log = tk.Text(sec3, height=7, font=("Consolas", 8), bg="#1A1A2E", fg="#00FF88",
                           relief="flat", state="disabled", wrap="word")
        self.log.pack(fill="x")
        sb = ttk.Scrollbar(sec3, command=self.log.yview)
        sb.pack(side="right", fill="y")
        self.log["yscrollcommand"] = sb.set

        # ── Bottoni ──
        btn_frame = tk.Frame(body, bg=BG)
        btn_frame.pack(fill="x")

        self.btn_run = tk.Button(btn_frame, text="Avvia elaborazione",
                                 font=("Calibri", 10, "bold"), bg=BLUE, fg=WHITE,
                                 activebackground=LBLUE, activeforeground=WHITE,
                                 relief="flat", padx=20, pady=8, cursor="hand2",
                                 command=self._run)
        self.btn_run.pack(side="left")

        self.btn_open = tk.Button(btn_frame, text="Apri output",
                                  font=("Calibri", 10), bg=GRAY, fg="#2D3748",
                                  relief="flat", padx=16, pady=8, cursor="hand2",
                                  state="disabled", command=self._open_output)
        self.btn_open.pack(side="left", padx=(10, 0))

        # ── Barra progresso ──
        self.progress = ttk.Progressbar(body, mode="indeterminate", length=400)
        self.progress.pack(pady=(12, 0))

    def _browse(self, var, filetypes):
        if "output" in str(var) or var == self.v_output:
            path = filedialog.asksaveasfilename(filetypes=filetypes,
                                                defaultextension=".xlsx")
        else:
            path = filedialog.askopenfilename(filetypes=filetypes)
        if path:
            var.set(path)

    def _log(self, msg, tag=None):
        self.log.configure(state="normal")
        self.log.insert("end", msg + "\n")
        self.log.see("end")
        self.log.configure(state="disabled")

    def _run(self):
        fat  = self.v_fatture.get().strip()
        csv_ = self.v_csv.get().strip()
        out  = self.v_output.get().strip()
        data = self.v_data.get().strip()
        sz   = self.v_sz.get().strip()

        if not fat or not csv_ or not out:
            messagebox.showwarning("Campi mancanti", "Seleziona tutti e tre i file prima di procedere.")
            return

        self.btn_run.configure(state="disabled")
        self.btn_open.configure(state="disabled")
        self.log.configure(state="normal"); self.log.delete("1.0", "end"); self.log.configure(state="disabled")
        self.progress.start(10)

        def worker():
            try:
                self._log(f"Carico FattureEmesse: {fat}")
                df_fat = carica_fatture(fat)
                self._log(f"   {len(df_fat)} righe caricate")

                self._log(f"Carico CSV SDI: {csv_}")
                df_csv = carica_csv(csv_)
                self._log(f"   {len(df_csv)} righe caricate")

                self._log("Elaboro incrocio...")
                result, totali = elabora(df_fat, df_csv, sz, data)

                ok   = (result["Stato"] == "✔ OK").sum()
                warn = (result["Stato"] != "✔ OK").sum()
                self._log(f"   Righe totali : {len(result)}")
                self._log(f"   OK           : {ok}")
                self._log(f"   Discrepanze  : {warn}")
                self._log(f"   Imponibile FattureEmesse : {totali['Imponibile FattureEmesse']:,.2f}")
                self._log(f"   Imponibile CSV           : {totali['Imponibile CSV']:,.2f}")
                self._log(f"   Delta imponibile         : {totali['Delta Imponibile']:,.2f}")
                if totali.get("Duplicati CSV", 0) > 0:
                    self._log(f"   Numeri fattura duplicati nel CSV: {totali['Duplicati CSV']} (tenuta la riga accettata)")

                self._log(f"Salvo output: {out}")
                Path(out).parent.mkdir(parents=True, exist_ok=True)
                formatta_excel(out, result, totali)
                self._log("Elaborazione completata.")
                self.btn_open.configure(state="normal")
            except Exception as e:
                self._log(f"ERRORE: {e}")
                messagebox.showerror("Errore", str(e))
            finally:
                self.progress.stop()
                self.btn_run.configure(state="normal")

        threading.Thread(target=worker, daemon=True).start()

    def _open_output(self):
        path = self.v_output.get().strip()
        if path and Path(path).exists():
            if sys.platform == "win32":
                os.startfile(path)
            elif sys.platform == "darwin":
                subprocess.call(["open", path])
            else:
                subprocess.call(["xdg-open", path])


if __name__ == "__main__":
    App().mainloop()